"""Extract per-question source material from the HF Modeloff dump.

Reads ``_hf_dsbench/data_analysis/data/<challenge_id>/*`` and writes:

    analysis/<slug>/source/introduction.md   — challenge intro
    analysis/<slug>/source/question_<N>.md   — per-question text (with A-I options)
    analysis/<slug>/source/excel_summary.md  — best-effort Excel/xlsx markdown summary
    analysis/<slug>/source/_manifest.json    — file inventory + extraction success

This is invoked ONCE off-line; downstream the LLM / heuristic ``_excel_agent``
reads only the per-question files. NEVER reads any answer key.

Per the protocol: this script extracts question wording + workbook context.
It NEVER touches ``_analysis_data.json`` answers.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sys
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
HF_DATA = ROOT / "_hf_dsbench" / "data_analysis" / "data"
ANALYSIS_DIR = ROOT / "analysis"
REGISTRY = ROOT / "registry" / "analysis_tasks.json"


def _read_text(p: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            return p.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return p.read_bytes().decode("utf-8", errors="replace")


def _excel_to_markdown(xlsx_path: Path, max_sheets: int = 6, max_rows_per_sheet: int = 80,
                       max_cols: int = 30) -> str:
    """Best-effort dump of the workbook to compact markdown the LLM can read.

    Returns "" if the file can't be opened (older .xlsb / .xlsm with macros
    sometimes fail under openpyxl)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as exc:
        return f"[could not open {xlsx_path.name}: {type(exc).__name__}: {exc}]"
    out: list[str] = [f"# Workbook: {xlsx_path.name}\n"]
    for sname in wb.sheetnames[:max_sheets]:
        try:
            ws = wb[sname]
            out.append(f"\n## Sheet: {sname}\n")
            rows_iter = ws.iter_rows(values_only=True)
            buf = []
            for r_idx, row in enumerate(rows_iter):
                if r_idx >= max_rows_per_sheet:
                    out.append(f"... ({r_idx}+ rows truncated)\n")
                    break
                clipped = row[:max_cols]
                # Skip fully-empty rows
                if all(c is None or (isinstance(c, str) and not c.strip()) for c in clipped):
                    continue
                vals = []
                for c in clipped:
                    if c is None:
                        vals.append("")
                    elif isinstance(c, float):
                        vals.append(f"{c:.6g}")
                    else:
                        vals.append(str(c).strip().replace("|", "\\|").replace("\n", " "))
                buf.append("| " + " | ".join(vals) + " |")
            if buf:
                out.append("\n".join(buf) + "\n")
        except Exception as exc:
            out.append(f"[sheet {sname} read error: {exc}]\n")
    return "".join(out)


def _challenge_dirs() -> dict[str, Path]:
    if not HF_DATA.exists():
        return {}
    out: dict[str, Path] = {}
    for d in HF_DATA.iterdir():
        if d.is_dir() and re.match(r"^\d{8}$", d.name):
            out[d.name] = d
    return out


def _slug_to_chid() -> dict[str, str]:
    reg = json.loads(REGISTRY.read_text(encoding="utf-8"))
    return {r["slug"]: r["challenge_id"] for r in reg}


def extract_one(slug: str, ch_dir: Path, q_names: list[str], out_dir: Path) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest: dict = {"slug": slug, "challenge_dir": str(ch_dir),
                       "extracted_questions": [], "extracted_intro": False,
                       "excel_files": [], "errors": []}
    # Introduction
    intro = ch_dir / "introduction.txt"
    if intro.exists():
        try:
            txt = _read_text(intro)
            (out_dir / "introduction.md").write_text(txt, encoding="utf-8")
            manifest["extracted_intro"] = True
            manifest["intro_chars"] = len(txt)
        except Exception as exc:
            manifest["errors"].append(f"intro: {exc}")
    # Per-question
    for q in q_names:
        qf = ch_dir / f"{q}.txt"
        if qf.exists():
            try:
                txt = _read_text(qf)
                # Question name → strip "questionN" prefix when writing
                m = re.match(r"^question(\d+)$", q)
                n = m.group(1) if m else q
                (out_dir / f"question_{n}.md").write_text(txt, encoding="utf-8")
                manifest["extracted_questions"].append({"name": q, "chars": len(txt)})
            except Exception as exc:
                manifest["errors"].append(f"{q}: {exc}")
        else:
            manifest["errors"].append(f"{q}: file not found")
    # Excel files (best-effort, may fail for .xlsb)
    for f in ch_dir.iterdir():
        if f.suffix.lower() in (".xlsx", ".xlsm"):
            try:
                md = _excel_to_markdown(f)
                if md:
                    safe = re.sub(r"[^A-Za-z0-9._-]", "_", f.name)
                    (out_dir / f"excel_{safe}.md").write_text(md, encoding="utf-8")
                    manifest["excel_files"].append({"file": f.name, "md_chars": len(md)})
            except Exception as exc:
                manifest["errors"].append(f"excel {f.name}: {exc}")
        elif f.suffix.lower() == ".xlsb":
            manifest["excel_files"].append({"file": f.name, "md_chars": 0,
                                            "note": "xlsb format — not parsed (no engine)"})
    (out_dir / "_manifest.json").write_text(json.dumps(manifest, indent=2),
                                             encoding="utf-8")
    return manifest


def main() -> None:
    if not HF_DATA.exists():
        print(f"FATAL: HF data dir not present at {HF_DATA}")
        sys.exit(1)
    ch_dirs = _challenge_dirs()
    print(f"[extract] {len(ch_dirs)} challenge dirs available in HF dump")
    slug_to_chid = _slug_to_chid()
    print(f"[extract] {len(slug_to_chid)} slugs in registry")
    # Load _analysis_data.json to get the question names per slug
    by_chid: dict[str, dict] = {}
    with (ROOT / "_analysis_data.json").open(encoding="utf-8-sig") as f:
        for line in f:
            if line.strip():
                row = json.loads(line)
                by_chid[row["id"]] = row
    summaries: list[dict] = []
    ok_full = 0
    ok_partial = 0
    failed = 0
    for slug, chid in slug_to_chid.items():
        repo = ANALYSIS_DIR / slug
        if not repo.exists():
            print(f"[extract] WARN: no repo for slug {slug}")
            failed += 1
            continue
        ch_dir = ch_dirs.get(chid)
        row = by_chid.get(chid)
        if ch_dir is None or row is None:
            print(f"[extract] WARN: no source dir / data row for chid {chid} (slug {slug})")
            failed += 1
            continue
        out_dir = repo / "source"
        try:
            m = extract_one(slug, ch_dir, row["questions"], out_dir)
            n_q = len(row["questions"])
            n_extracted = len(m["extracted_questions"])
            status = "FULL" if (n_extracted == n_q and m["extracted_intro"]) else "PARTIAL"
            print(f"[extract] {slug}: {status}  ({n_extracted}/{n_q} questions, "
                  f"intro={m['extracted_intro']}, excel={len(m['excel_files'])} files)")
            summaries.append({"slug": slug, "chid": chid,
                               "n_questions": n_q,
                               "n_extracted_questions": n_extracted,
                               "intro": m["extracted_intro"],
                               "excel_md_files": len([e for e in m["excel_files"] if e.get("md_chars", 0) > 0]),
                               "excel_xlsb_skipped": len([e for e in m["excel_files"] if e.get("note", "").startswith("xlsb")]),
                               "status": status})
            if status == "FULL":
                ok_full += 1
            else:
                ok_partial += 1
        except Exception as exc:
            tb = traceback.format_exc()
            print(f"[extract] {slug}: FAILED  {type(exc).__name__}: {exc}")
            summaries.append({"slug": slug, "chid": chid, "error": str(exc), "tb": tb})
            failed += 1
    out_path = ROOT / "registry" / "modeloff_source_manifest.json"
    out_path.write_text(json.dumps(summaries, indent=2), encoding="utf-8")
    print(f"\n[extract] FULL={ok_full}  PARTIAL={ok_partial}  FAILED={failed}  / 38")
    print(f"[extract] manifest written to {out_path}")


if __name__ == "__main__":
    main()
